"""Pure spreadsheet-to-narrative conversion logic for the Excel Narrator Agent.

This module intentionally has ZERO a2a-sdk imports and ZERO imports from the
main project's `config` or `src` packages. It is plain, unit-testable Python
that turns spreadsheet bytes into descriptive natural-language paragraphs
suitable for embedding and retrieval in a RAG pipeline. The A2A server
(`services/excel_agent/server.py`) is the only piece of this package that
knows about the A2A protocol; it calls into this module for the actual work.

Deterministic baseline (always on): for each sheet, emit a context header
paragraph describing the sheet name, source filename, row count and column
names, followed by one sentence per data row (the row's first column is
treated as its "subject"). Optionally, when Azure AI Foundry GPT-4 creds are
configured, a 2-3 sentence LLM-generated summary is inserted between the
header paragraph and the row sentences. The LLM step is best-effort: any
failure (missing creds, network error, bad response, ...) falls back to the
deterministic output, so this agent always works fully offline.
"""

from __future__ import annotations

import csv
import io
import logging
import os
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openai import AzureOpenAI

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}

# Environment variables consulted for the optional per-sheet LLM summary.
# Read directly from the process environment (never from the main project's
# `config` package) to keep the two codebases decoupled.
_LLM_ENV_VARS = (
    "AZURE_AI_FOUNDRY_GPT4_ENDPOINT",
    "AZURE_AI_FOUNDRY_GPT4_DEPLOYMENT_NAME",
    "AZURE_AI_FOUNDRY_GPT4_API_KEY",
    "AZURE_AI_FOUNDRY_GPT4_API_VERSION",
)


class NarratorError(Exception):
    """Raised when a spreadsheet cannot be parsed into sheets/rows."""


@dataclass
class ParsedSheet:
    """One sheet's headers and (already blank-row-filtered) data rows."""

    name: str
    headers: list[str] = field(default_factory=list)
    data_rows: list[list] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    """A spreadsheet parsed into sheets, before any narrative text is built."""

    filename: str
    sheets: list[ParsedSheet] = field(default_factory=list)

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.name for sheet in self.sheets]

    @property
    def row_count(self) -> int:
        return sum(len(sheet.data_rows) for sheet in self.sheets)


def _format_value(value: object) -> str | None:
    """Format one cell value for narrative text, or None if the cell is empty."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, bool):
        # bool is a subclass of int, so this must be checked before int/float.
        return "Yes" if value else "No"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _row_is_empty(values: list) -> bool:
    return all(_format_value(v) is None for v in values)


def _format_header(value: object, index: int) -> str:
    formatted = _format_value(value)
    return formatted if formatted is not None else f"Column{index + 1}"


def _sheet_header_paragraph(
    sheet_name: str, filename: str, data_row_count: int, headers: list[str]
) -> str:
    columns = ", ".join(headers)
    row_word = "row" if data_row_count == 1 else "rows"
    return (
        f"Sheet '{sheet_name}' from {filename} contains {data_row_count} data "
        f"{row_word} with columns: {columns}."
    )


def _narrate_row(sheet_name: str, headers: list[str], row: list) -> str:
    subject_header = headers[0] if headers else "Row"
    subject_value = _format_value(row[0]) if row else None
    if subject_value is None:
        subject_value = "(unspecified)"

    clauses = []
    for header, value in zip(headers[1:], row[1:]):
        formatted = _format_value(value)
        if formatted is None:
            continue
        clauses.append(f"{header} is {formatted}")

    return f"{sheet_name} — {subject_header} '{subject_value}': {'; '.join(clauses)}."


class ExcelNarrator:
    """Converts spreadsheet bytes into descriptive natural-language passages.

    The API is split into `load()` (fast, parses the file into sheets/rows)
    and `narrate_sheet()` (builds the narrative text for a single
    already-parsed sheet, including the optional LLM summary). Splitting it
    this way lets a caller (e.g. the A2A server) emit a progress update for
    sheet N *before* doing that sheet's work -- not possible if narration
    happened as one big blocking call. `narrate()` and `iter_sheets()` are
    convenience wrappers for callers that just want the end result.
    """

    def __init__(self) -> None:
        self.use_llm = self._llm_enabled()
        self._client: AzureOpenAI | None = None
        if self.use_llm:
            logger.info(
                "ExcelNarrator: per-sheet LLM summaries ENABLED "
                "(Azure AI Foundry GPT-4 deployment)."
            )
        else:
            logger.info(
                "ExcelNarrator: per-sheet LLM summaries DISABLED "
                "(deterministic templates only)."
            )

    @staticmethod
    def _llm_enabled() -> bool:
        if os.environ.get("EXCEL_AGENT_USE_LLM", "true").strip().lower() == "false":
            return False
        return all(os.environ.get(name) for name in _LLM_ENV_VARS)

    # --- Parsing -------------------------------------------------------

    def load(self, data: bytes, filename: str) -> ParsedWorkbook:
        """Parse spreadsheet bytes into sheets and rows. No narrative text yet."""
        suffix = Path(filename).suffix.lower()
        if suffix == ".xlsx":
            raw_sheets = self._read_xlsx(data)
        elif suffix == ".csv":
            raw_sheets = [(Path(filename).stem, self._read_csv(data))]
        else:
            raise NarratorError(
                f"Unsupported file extension '{suffix}'. Expected one of: "
                f"{', '.join(sorted(SUPPORTED_EXTENSIONS))}."
            )

        sheets = [self._build_sheet(name, rows) for name, rows in raw_sheets]
        return ParsedWorkbook(filename=filename, sheets=sheets)

    @staticmethod
    def _read_xlsx(data: bytes) -> list[tuple[str, list[list]]]:
        workbook = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True, read_only=True
        )
        try:
            return [
                (ws.title, [list(row) for row in ws.iter_rows(values_only=True)])
                for ws in workbook.worksheets
            ]
        finally:
            workbook.close()

    @staticmethod
    def _read_csv(data: bytes) -> list[list]:
        text = data.decode("utf-8-sig")
        return [list(row) for row in csv.reader(io.StringIO(text))]

    @staticmethod
    def _build_sheet(sheet_name: str, rows: list[list]) -> ParsedSheet:
        if not rows:
            return ParsedSheet(name=sheet_name)
        header_row = rows[0]
        headers = [_format_header(h, i) for i, h in enumerate(header_row)]
        data_rows = [row for row in rows[1:] if not _row_is_empty(row)]
        return ParsedSheet(name=sheet_name, headers=headers, data_rows=data_rows)

    # --- Narration -------------------------------------------------------

    def narrate_sheet(self, workbook: ParsedWorkbook, sheet: ParsedSheet) -> str:
        """Build the full narrative text (header + optional summary + rows) for one sheet."""
        header_paragraph = _sheet_header_paragraph(
            sheet.name, workbook.filename, len(sheet.data_rows), sheet.headers
        )
        row_sentences = [
            _narrate_row(sheet.name, sheet.headers, row) for row in sheet.data_rows
        ]

        paragraphs = [header_paragraph]

        if self.use_llm:
            summary = self._summarize_sheet(sheet.name, header_paragraph, row_sentences)
            if summary:
                paragraphs.append(summary)

        if row_sentences:
            paragraphs.append("\n".join(row_sentences))

        return "\n\n".join(paragraphs)

    def iter_sheets(self, data: bytes, filename: str) -> Iterator[tuple[str, str]]:
        """Convenience generator yielding (sheet_name, sheet_text) per sheet.

        Parses then narrates lazily, one sheet at a time. The A2A server uses
        `load()` + `narrate_sheet()` directly instead, so it can publish a
        "working on sheet N" status update *before* that sheet's (possibly
        LLM-backed) work runs; this generator is here for simpler callers.
        """
        workbook = self.load(data, filename)
        for sheet in workbook.sheets:
            yield sheet.name, self.narrate_sheet(workbook, sheet)

    def narrate(self, data: bytes, filename: str) -> dict:
        """Parse and narrate every sheet, returning the full result at once."""
        workbook = self.load(data, filename)
        sheet_texts = [self.narrate_sheet(workbook, sheet) for sheet in workbook.sheets]
        return {
            "text": "\n\n".join(sheet_texts),
            "sheet_count": len(workbook.sheets),
            "sheet_names": workbook.sheet_names,
            "row_count": workbook.row_count,
        }

    # --- Optional LLM summary -------------------------------------------

    def _get_client(self) -> AzureOpenAI:
        if self._client is None:
            self._client = AzureOpenAI(
                azure_endpoint=os.environ["AZURE_AI_FOUNDRY_GPT4_ENDPOINT"],
                api_key=os.environ["AZURE_AI_FOUNDRY_GPT4_API_KEY"],
                api_version=os.environ["AZURE_AI_FOUNDRY_GPT4_API_VERSION"],
            )
        return self._client

    def _summarize_sheet(
        self, sheet_name: str, header_paragraph: str, row_sentences: list[str]
    ) -> str | None:
        """One best-effort LLM call for a 2-3 sentence factual sheet summary.

        Any exception (missing creds, network failure, bad response, ...) is
        logged as a warning and swallowed -- the agent must keep working
        fully offline even when the LLM is unreachable.
        """
        try:
            client = self._get_client()
            sample = "\n".join(row_sentences[:15])
            prompt = f"{header_paragraph}\n\n{sample}"
            response = client.chat.completions.create(
                model=os.environ["AZURE_AI_FOUNDRY_GPT4_DEPLOYMENT_NAME"],
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a precise data analyst. Given a spreadsheet "
                            "sheet's description and a sample of its rows, write a "
                            "factual 2-3 sentence summary of what the table shows. "
                            "State only facts supported by the given rows."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
            )
            summary = response.choices[0].message.content
            return summary.strip() if summary else None
        except Exception as e:  # noqa: BLE001 - best-effort, must never break narration
            logger.warning(
                f"LLM summary failed for sheet '{sheet_name}', continuing "
                f"without it: {e}"
            )
            return None
